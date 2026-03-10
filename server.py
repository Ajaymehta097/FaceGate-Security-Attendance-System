"""
FaceGate Web Server — Updated Version
======================================================
Changes:
  1. Admin sirf EK BAAR register hoga (pehli baar aap khud karein)
     - Iske baad /admin/register tab lock ho jaata hai
  2. Admin face scan se login hoga → Admin page khulega
  3. Student / Staff / Admin ka data ALAG ALAG
  4. Excel export mein role-wise alag sheets
  5. Web UI mein bhi role-wise filter
  6. Unknown person → sound alert + screenshot save
  7. Better registration flow
"""

from flask import Flask, Response, jsonify, request, send_file
from flask_cors import CORS
import cv2
import face_recognition
import numpy as np
import base64
import json
import os
import time
import datetime
import pickle
import threading
import secrets

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False


try:
    import qrcode
    from PIL import Image as PILImage
    QR_OK = True
except ImportError:
    QR_OK = False

# OpenCV QR decoder — no extra library needed
QR_DETECTOR = cv2.QRCodeDetector()

app = Flask(__name__)
CORS(app)

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
DB_FILE         = os.path.join(BASE_DIR, "data", "face_database.pkl")
LOG_FILE        = os.path.join(BASE_DIR, "logs", "access_log.json")
KNOWN_DIR       = os.path.join(BASE_DIR, "known_faces")
ADMIN_DB_FILE   = os.path.join(BASE_DIR, "data", "admin.pkl")
ATTENDANCE_FILE = os.path.join(BASE_DIR, "data", "attendance.json")
ALERTS_DIR      = os.path.join(BASE_DIR, "alerts")
SETTINGS_FILE   = os.path.join(BASE_DIR, "data", "settings.json")

# Default settings
TOLERANCE    = 0.5
FRAME_SCALE  = 0.5
CAMERA_INDEX = 0
SCAN_INTERVAL = 1.5

os.makedirs(os.path.join(BASE_DIR, "data"), exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "logs"), exist_ok=True)
os.makedirs(KNOWN_DIR, exist_ok=True)
os.makedirs(ALERTS_DIR, exist_ok=True)
QR_DIR = os.path.join(BASE_DIR, "data", "qrcodes")
os.makedirs(QR_DIR, exist_ok=True)

# ── Shared camera state ───────────────────────────────────────────────────────
camera_active  = False
cap            = None
camera_lock    = threading.Lock()
last_result    = {"status": "idle", "name": "", "confidence": 0, "role": ""}
result_until   = 0
cooldown_until = 0
admin_sessions = set()

@app.route("/")
def home():
    return "FaceGate Running Successfully"
# ── Settings load/save ────────────────────────────────────────────────────────
def load_settings():
    global TOLERANCE, CAMERA_INDEX, SCAN_INTERVAL
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE) as f:
            s = json.load(f)
        TOLERANCE     = s.get("tolerance", 0.5)
        CAMERA_INDEX  = s.get("camera_index", 0)
        SCAN_INTERVAL = s.get("scan_interval", 1.5)

def save_settings_file():
    with open(SETTINGS_FILE, "w") as f:
        json.dump({"tolerance": TOLERANCE, "camera_index": CAMERA_INDEX,
                   "scan_interval": SCAN_INTERVAL}, f)

load_settings()

# ── Admin DB ─────────────────────────────────────────────────────────────────
def load_admin_db():
    if os.path.exists(ADMIN_DB_FILE):
        with open(ADMIN_DB_FILE, "rb") as f:
            return pickle.load(f)
    return []

def save_admin_db(admins):
    with open(ADMIN_DB_FILE, "wb") as f:
        pickle.dump(admins, f)

def admin_registered():
    """Returns True if at least one admin is registered."""
    return len(load_admin_db()) > 0

def is_admin_session(req):
    token = req.headers.get("X-Admin-Token", "") or req.args.get("token", "")
    return token in admin_sessions

# ── Student DB ────────────────────────────────────────────────────────────────
def load_db():
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "rb") as f:
            return pickle.load(f)
    return []

def save_db(records):
    os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
    with open(DB_FILE, "wb") as f:
        pickle.dump(records, f)

def load_db_by_role(role=None):
    """Load all records or filter by role (student/staff/admin)."""
    records = load_db()
    if role:
        return [r for r in records if r.get("role","").lower() == role.lower()]
    return records

# ── Logs ─────────────────────────────────────────────────────────────────────
def load_logs(n=100):
    if not os.path.exists(LOG_FILE):
        return []
    with open(LOG_FILE, "r") as f:
        logs = json.load(f)
    return logs[-n:]

def append_log(entry):
    logs = []
    if os.path.exists(LOG_FILE):
        try:
            with open(LOG_FILE, "r") as f:
                logs = json.load(f)
        except:
            logs = []
    logs.append(entry)
    with open(LOG_FILE, "w") as f:
        json.dump(logs, f, indent=2)

# ── Attendance ────────────────────────────────────────────────────────────────
def load_attendance():
    if os.path.exists(ATTENDANCE_FILE):
        with open(ATTENDANCE_FILE, "r") as f:
            return json.load(f)
    return {}

def mark_attendance(student_id, name, role, dept):
    att   = load_attendance()
    today = datetime.date.today().isoformat()
    now_t = datetime.datetime.now().strftime("%H:%M:%S")
    if student_id not in att:
        att[student_id] = {"name": name, "role": role, "dept": dept, "records": {}}
    if today not in att[student_id]["records"]:
        att[student_id]["records"][today] = now_t
        with open(ATTENDANCE_FILE, "w") as f:
            json.dump(att, f, indent=2)
        return True
    return False

def get_attendance_percent(student_id):
    att = load_attendance()
    if student_id not in att:
        return 0
    recs = att[student_id]["records"]
    total_days = max((datetime.date.today() - datetime.date(2025, 1, 1)).days, 1)
    present = len(recs)
    return round((present / min(present + 10, total_days)) * 100, 1)

# ── Alert screenshot ──────────────────────────────────────────────────────────
last_alert_time = 0

def save_unknown_alert(frame):
    global last_alert_time
    now = time.time()
    if now - last_alert_time < 10:  # max 1 alert per 10 sec
        return
    last_alert_time = now
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(ALERTS_DIR, f"unknown_{ts}.jpg")
    cv2.imwrite(path, frame)
    alerts_log = os.path.join(BASE_DIR, "data", "alerts.json")
    alerts = []
    if os.path.exists(alerts_log):
        try:
            with open(alerts_log) as f:
                alerts = json.load(f)
        except:
            pass
    alerts.append({
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "img": path
    })
    with open(alerts_log, "w") as f:
        json.dump(alerts[-50:], f, indent=2)

# ── Image helper ─────────────────────────────────────────────────────────────
def _img_b64(path):
    if path and os.path.exists(path):
        with open(path, "rb") as f:
            return "data:image/jpeg;base64," + base64.b64encode(f.read()).decode()
    return ""


def generate_qr(roll_id, name, role, dept):
    """Generate QR code for a student/staff — stores in data/qrcodes/"""
    if not QR_OK:
        return ""
    qr_data = f"FACEGATE|{roll_id}|{name}|{role}|{dept}"
    qr_path = os.path.join(QR_DIR, f"{roll_id}.png")
    img = qrcode.make(qr_data)
    img.save(qr_path)
    return qr_path

def _qr_b64(roll_id):
    path = os.path.join(QR_DIR, f"{roll_id}.png")
    return _img_b64(path)

# ── Camera frame generator (MJPEG) ───────────────────────────────────────────
def generate_frames():
    global cap, camera_active, last_result, result_until, cooldown_until

    # Load students/staff from main DB
    records = load_db()

    # ✅ FIX: Admin bhi scanner mein enter ho sake —
    # admin.pkl se admins load karo aur unhe bhi list mein add karo
    admins = load_admin_db()
    admin_ids_in_db = {r["id"] for r in records if r.get("role") == "admin"}
    for adm in admins:
        adm_id = adm.get("id", f"ADMIN_{adm['name'].replace(' ','_').upper()}")
        if adm_id not in admin_ids_in_db:
            # Admin face_database mein nahi hai — dynamically add karo
            records.append({
                "name"    : adm["name"],
                "id"      : adm_id,
                "role"    : "admin",
                "dept"    : "Administration",
                "encoding": adm["encoding"],
                "img_path": "",
            })

    known_encs  = [r["encoding"]        for r in records]
    known_names = [r["name"]            for r in records]
    known_ids   = [r["id"]              for r in records]
    known_roles = [r["role"]            for r in records]
    known_depts = [r.get("dept","N/A")  for r in records]
    last_scan_time = 0

    cap = cv2.VideoCapture(CAMERA_INDEX)
    camera_active = True

    try:
        while camera_active:
            ret, frame = cap.read()
            if not ret:
                break

            frame = cv2.flip(frame, 1)
            now   = time.time()

            small = cv2.resize(frame, (0, 0), fx=FRAME_SCALE, fy=FRAME_SCALE)
            rgb   = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
            locs  = face_recognition.face_locations(rgb)
            encs  = face_recognition.face_encodings(rgb, locs)

            if now - last_scan_time > SCAN_INTERVAL and encs and now > cooldown_until:
                last_scan_time = now
                for enc in encs:
                    distances = face_recognition.face_distance(known_encs, enc)
                    best      = int(np.argmin(distances)) if len(distances) > 0 else -1

                    if best >= 0 and distances[best] < TOLERANCE:
                        conf  = round((1 - distances[best]) * 100, 1)
                        dept  = known_depts[best]
                        role  = known_roles[best]
                        newly = mark_attendance(known_ids[best], known_names[best], role, dept)
                        att_pct = get_attendance_percent(known_ids[best])

                        last_result = {
                            "status"      : "granted",
                            "name"        : known_names[best],
                            "id"          : known_ids[best],
                            "role"        : role,
                            "dept"        : dept,
                            "confidence"  : conf,
                            "att_pct"     : att_pct,
                            "newly_marked": newly,
                            "is_admin"    : role.lower() == "admin",
                            "time"        : datetime.datetime.now().strftime("%H:%M:%S"),
                        }
                        append_log({
                            **{k: v for k, v in last_result.items() if k != "is_admin"},
                            "timestamp"  : datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "status_label": "GRANTED"
                        })
                    else:
                        last_result = {"status": "denied", "name": "UNKNOWN",
                                       "confidence": 0, "role": "", "dept": "", "att_pct": 0, "is_admin": False}
                        append_log({
                            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "name": "UNKNOWN", "id": "—", "role": "—",
                            "status_label": "DENIED", "confidence": 0
                        })
                        threading.Thread(target=save_unknown_alert,
                                         args=(frame.copy(),), daemon=True).start()

                    result_until   = now + 4
                    cooldown_until = now + 5
                    break

            # Draw face boxes
            scale = int(1 / FRAME_SCALE)
            for (top, right, bottom, left) in locs:
                t, r, b, l = top*scale, right*scale, bottom*scale, left*scale
                clr = (185, 80, 255) if (last_result.get("status") == "granted" and now < result_until) \
                    else (60, 40, 220) if (last_result.get("status") == "denied" and now < result_until) \
                    else (180, 180, 0)
                cv2.rectangle(frame, (l, t), (r, b), clr, 2)

            # Big overlay
            if last_result and now < result_until:
                h, w  = frame.shape[:2]
                overlay = frame.copy()

                if last_result["status"] == "granted":
                    card_h = 170
                    cv2.rectangle(overlay, (0, h-card_h), (w, h), (30, 10, 60), -1)
                    cv2.addWeighted(overlay, 0.82, frame, 0.18, 0, frame)
                    cv2.rectangle(frame, (0, h-card_h), (w, h-card_h+3), (185, 80, 255), -1)

                    role_str = last_result["role"].upper()
                    enter_txt = f"✓  YOU CAN ENTER  [{role_str}]"
                    cv2.putText(frame, enter_txt, (20, h-card_h+30),
                                cv2.FONT_HERSHEY_DUPLEX, 0.85, (185, 80, 255), 2)

                    cv2.putText(frame, f"Name       : {last_result['name']}",
                                (20, h-card_h+60), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 1)
                    cv2.putText(frame, f"ID         : {last_result['id']}",
                                (20, h-card_h+84), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 1)
                    cv2.putText(frame, f"Department : {last_result.get('dept','N/A')}",
                                (20, h-card_h+108), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 1)
                    att_note = "  ✓ Marked Today" if last_result.get("newly_marked") else "  (Already Marked)"
                    cv2.putText(frame, f"Attendance : {last_result.get('att_pct',0)}%" + att_note,
                                (20, h-card_h+132), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (80, 255, 150), 1)

                    bar_w = int((w - 40) * last_result.get("att_pct", 0) / 100)
                    cv2.rectangle(frame, (20, h-18), (w-20, h-8), (60, 30, 90), -1)
                    cv2.rectangle(frame, (20, h-18), (20+bar_w, h-8), (185, 80, 255), -1)

                    # If admin — show special badge
                    if last_result.get("is_admin"):
                        cv2.putText(frame, "[ ADMIN — Full Access ]",
                                    (w-280, h-card_h+30), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 191, 36), 2)
                else:
                    cv2.rectangle(overlay, (0, h-90), (w, h), (20, 10, 100), -1)
                    cv2.addWeighted(overlay, 0.85, frame, 0.15, 0, frame)
                    cv2.rectangle(frame, (0, h-90), (w, h-87), (60, 40, 220), -1)
                    cv2.putText(frame, "!! YOU CANNOT ENTER !!",
                                (20, h-58), cv2.FONT_HERSHEY_DUPLEX, 0.85, (100, 80, 255), 2)
                    cv2.putText(frame, "Not registered — Alert saved to admin",
                                (20, h-22), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (200, 180, 255), 1)

            ts = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
            cv2.putText(frame, f"FACEGATE  |  {ts}", (10, 24),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (185, 80, 255), 1)

            _, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
    finally:
        if cap:
            cap.release()
        camera_active = False


# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    with open(os.path.join(BASE_DIR, "templates", "index.html"), encoding="utf-8") as f:
        return f.read()

@app.route("/admin")
def admin_page():
    with open(os.path.join(BASE_DIR, "templates", "admin.html"), encoding="utf-8") as f:
        return f.read()

@app.route("/video_feed")
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route("/stop_camera", methods=["POST"])
def stop_camera():
    global camera_active
    camera_active = False
    return jsonify({"ok": True})

@app.route("/status")
def status():
    now = time.time()
    if now < result_until:
        return jsonify(last_result)
    return jsonify({"status": "idle"})

# ── Registration ──────────────────────────────────────────────────────────────

@app.route("/register", methods=["POST"])
def register():
    data    = request.json
    name    = data.get("name", "").strip()
    roll_id = data.get("id", "").strip()
    role    = data.get("role", "student").strip().lower()
    dept    = data.get("dept", "General").strip()
    img_b64 = data.get("image", "")

    if not name or not roll_id or not img_b64:
        return jsonify({"ok": False, "error": "Name, ID, aur photo required hain."})

    # Admin role register sirf tab allowed agar admin already registered nahi
    # (Admin panel se naye admins add ho sakte hain — sirf seedha register block)
    if role == "admin":
        if admin_registered():
            return jsonify({"ok": False,
                            "error": "Admin pehle se registered hai. Naya admin sirf Admin Panel se add ho sakta hai."})

    img_data = base64.b64decode(img_b64.split(",")[-1])
    nparr    = np.frombuffer(img_data, np.uint8)
    frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    rgb      = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    locs = face_recognition.face_locations(rgb)
    encs = face_recognition.face_encodings(rgb, locs)

    if not encs:
        return jsonify({"ok": False, "error": "Photo mein koi face nahi mila. Better lighting mein try karein."})

    img_path = os.path.join(KNOWN_DIR, f"{role}_{roll_id}_{name.replace(' ','_')}.jpg")
    cv2.imwrite(img_path, frame)

    records = load_db()
    records = [r for r in records if r["id"] != roll_id]
    records.append({
        "name": name, "id": roll_id, "role": role,
        "dept": dept, "encoding": encs[0], "img_path": img_path
    })
    save_db(records)

    # Generate QR code for this person
    qr_path = generate_qr(roll_id, name, role, dept)

    # If admin registered via this route — also save to admin DB
    if role == "admin":
        admins = load_admin_db()
        admins = [a for a in admins if a.get("id") != roll_id]
        admins.append({"name": name, "id": roll_id, "encoding": encs[0]})
        save_admin_db(admins)

    return jsonify({"ok": True, "message": f"{name} ({role.upper()}) successfully registered!", "qr": _qr_b64(roll_id), "id": roll_id})

@app.route("/admin_registered")
def check_admin_registered():
    """Frontend checks this to show/hide admin register tab."""
    return jsonify({"registered": admin_registered(), "count": len(load_admin_db())})

# ── Students (role-filtered) ──────────────────────────────────────────────────

@app.route("/students")
def get_students():
    role = request.args.get("role", None)   # ?role=student / staff / admin
    records = load_db_by_role(role)
    safe = [{
        "name": r["name"], "id": r["id"],
        "role": r["role"], "dept": r.get("dept","N/A"),
        "img": _img_b64(r.get("img_path","")),
        "qr": _qr_b64(r["id"])
    } for r in records]
    return jsonify(safe)

@app.route("/logs")
def get_logs():
    return jsonify(load_logs(100))

@app.route("/delete/<roll_id>", methods=["DELETE"])
def delete_student(roll_id):
    records = load_db()
    records = [r for r in records if r["id"] != roll_id]
    save_db(records)
    return jsonify({"ok": True})

# ── Attendance ────────────────────────────────────────────────────────────────

@app.route("/attendance")
def get_attendance():
    role    = request.args.get("role", None)
    att     = load_attendance()
    records = load_db_by_role(role)
    today   = datetime.date.today().isoformat()
    result  = []
    for r in records:
        sid  = r["id"]
        recs = att.get(sid, {}).get("records", {})
        pct  = get_attendance_percent(sid)
        result.append({
            "name"   : r["name"],
            "id"     : sid,
            "dept"   : r.get("dept", "N/A"),
            "role"   : r["role"],
            "present": len(recs),
            "today"  : today in recs,
            "pct"    : pct,
            "last"   : max(recs.keys()) if recs else "Never",
            "img"    : _img_b64(r.get("img_path", "")),
        })
    return jsonify(result)

@app.route("/attendance/export")
def export_attendance():
    """Export Excel with separate sheets for Student, Staff, Admin."""
    if not EXCEL_OK:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    att     = load_attendance()
    records = load_db()
    today   = datetime.date.today().isoformat()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    roles = ["student", "staff", "admin"]
    for role in roles:
        role_records = [r for r in records if r.get("role","").lower() == role]
        if not role_records:
            continue

        ws = wb.create_sheet(title=role.capitalize())

        # Title
        ws.merge_cells("A1:I1")
        ws["A1"] = f"FaceGate — {role.upper()} Attendance Report ({today})"
        ws["A1"].font      = Font(bold=True, color="B94FFF", size=13)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28

        # Header
        hdr_fill = PatternFill("solid", fgColor="2D0A4E")
        hdr_font = Font(bold=True, color="E9D5FF", size=10)
        headers  = ["#", "Name", "ID / Roll No", "Department", "Role",
                    "Present Days", "Attendance %", "Today Present", "Last Seen"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

        col_widths = [4, 22, 14, 18, 10, 13, 13, 13, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64+i)].width = w

        grn_font = Font(color="34D399", bold=True)
        red_font = Font(color="F43F5E", bold=True)
        alt_fill = PatternFill("solid", fgColor="160F26")

        for i, r in enumerate(role_records, 1):
            sid   = r["id"]
            recs  = att.get(sid, {}).get("records", {})
            pct   = get_attendance_percent(sid)
            today_p = "YES" if today in recs else "NO"
            last_s  = max(recs.keys()) if recs else "Never"
            row = [i, r["name"], sid, r.get("dept","N/A"), r["role"],
                   len(recs), f"{pct}%", today_p, last_s]
            for col, val in enumerate(row, 1):
                c = ws.cell(row=i+2, column=col, value=val)
                c.alignment = Alignment(horizontal="center")
                if i % 2 == 0: c.fill = alt_fill
                if col == 7:
                    c.font = grn_font if pct >= 75 else red_font
                if col == 8:
                    c.font = grn_font if today_p == "YES" else red_font

    path = os.path.join(BASE_DIR, "data", f"attendance_{today}.xlsx")
    wb.save(path)
    return send_file(path, as_attachment=True,
                     download_name=f"FaceGate_Attendance_{today}.xlsx")

# ── Alerts ────────────────────────────────────────────────────────────────────

@app.route("/alerts")
def get_alerts():
    alerts_log = os.path.join(BASE_DIR, "data", "alerts.json")
    if not os.path.exists(alerts_log):
        return jsonify([])
    with open(alerts_log) as f:
        alerts = json.load(f)
    result = []
    for a in reversed(alerts[-20:]):
        result.append({
            "timestamp": a["timestamp"],
            "img": _img_b64(a.get("img",""))
        })
    return jsonify(result)

# ── Profile ───────────────────────────────────────────────────────────────────

@app.route("/profile/<roll_id>")
def get_profile(roll_id):
    records = load_db()
    r = next((x for x in records if x["id"] == roll_id), None)
    if not r:
        return jsonify({"error": "Not found"}), 404
    pct  = get_attendance_percent(roll_id)
    att  = load_attendance()
    recs = att.get(roll_id, {}).get("records", {})
    return jsonify({
        "name"   : r["name"],     "id":   roll_id,
        "dept"   : r.get("dept","N/A"), "role": r["role"],
        "pct"    : pct,           "present": len(recs),
        "img"    : _img_b64(r.get("img_path","")),
        "last"   : max(recs.keys()) if recs else "Never",
    })

# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/check")
def admin_check():
    return jsonify({"ok": is_admin_session(request)})

@app.route("/admin/register_admin", methods=["POST"])
def register_admin_route():
    """
    First-time admin registration (no token needed).
    After first admin is set, only existing admin can add more via /admin panel.
    """
    data    = request.json
    name    = data.get("name", "").strip()
    img_b64 = data.get("image", "")

    if not name or not img_b64:
        return jsonify({"ok": False, "error": "Name aur photo required hai."})

    # Check: agar already ek admin registered hai aur ye request admin panel se nahi aayi
    already_admin = admin_registered()
    is_admin_req  = is_admin_session(request)

    if already_admin and not is_admin_req:
        return jsonify({"ok": False,
                        "error": "Admin pehle se register hai. Naya admin sirf Admin Panel se register hoga."})

    img_data = base64.b64decode(img_b64.split(",")[-1])
    nparr    = np.frombuffer(img_data, np.uint8)
    frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    rgb      = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    encs     = face_recognition.face_encodings(rgb, face_recognition.face_locations(rgb))

    if not encs:
        return jsonify({"ok": False, "error": "Photo mein face nahi mila. Dobara try karein."})

    admin_id = f"ADMIN_{name.replace(' ','_').upper()}"

    # Save to admin.pkl WITH consistent id — scanner duplicate avoid karega
    admins = load_admin_db()
    admins = [a for a in admins if a.get("name") != name]
    admins.append({"name": name, "id": admin_id, "encoding": encs[0]})
    save_admin_db(admins)

    # ✅ Save to face_database.pkl — scanner yahan se match karta hai
    img_path = os.path.join(KNOWN_DIR, f"admin_{admin_id}.jpg")
    cv2.imwrite(img_path, frame)
    records = load_db()
    records = [r for r in records if r.get("id") != admin_id]
    records.append({
        "name": name, "id": admin_id,
        "role": "admin", "dept": "Administration",
        "encoding": encs[0], "img_path": img_path
    })
    save_db(records)

    return jsonify({"ok": True, "message": f"{name} admin ke roop mein register ho gaye!"})

@app.route("/admin/login", methods=["POST"])
def admin_login():
    data    = request.json
    img_b64 = data.get("image", "")
    if not img_b64:
        return jsonify({"ok": False, "error": "Photo required."})

    img_data = base64.b64decode(img_b64.split(",")[-1])
    nparr    = np.frombuffer(img_data, np.uint8)
    frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    rgb      = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    encs     = face_recognition.face_encodings(rgb, face_recognition.face_locations(rgb))

    if not encs:
        return jsonify({"ok": False, "error": "Koi face detect nahi hua."})

    admins = load_admin_db()
    if not admins:
        return jsonify({"ok": False, "error": "Koi admin register nahi hai abhi tak."})

    known = [a["encoding"] for a in admins]
    dists = face_recognition.face_distance(known, encs[0])
    best  = int(np.argmin(dists))

    if dists[best] < TOLERANCE:
        token = secrets.token_hex(16)
        admin_sessions.add(token)
        return jsonify({"ok": True, "token": token, "name": admins[best]["name"]})

    return jsonify({"ok": False, "error": "Aapka chehra admin ke roop mein nahi pehchana gaya."})

@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    token = (request.json or {}).get("token", "")
    admin_sessions.discard(token)
    return jsonify({"ok": True})

@app.route("/admin/students")
def admin_students():
    if not is_admin_session(request):
        return jsonify({"error": "Unauthorized"}), 401
    role    = request.args.get("role", None)
    records = load_db_by_role(role)
    today   = datetime.date.today().isoformat()
    logs    = load_logs(500)
    result  = []
    for r in records:
        s_logs      = [l for l in logs if l.get("id") == r["id"]]
        last_seen   = s_logs[-1]["timestamp"] if s_logs else "Never"
        today_count = sum(1 for l in s_logs if l.get("timestamp","").startswith(today))
        result.append({
            "name"        : r["name"], "id": r["id"],
            "role"        : r["role"], "dept": r.get("dept","N/A"),
            "img"         : _img_b64(r.get("img_path","")),
            "last_seen"   : last_seen,
            "today"       : today_count,
            "total_visits": len(s_logs),
        })
    return jsonify(result)

@app.route("/admin/stats")
def admin_stats():
    if not is_admin_session(request):
        return jsonify({"error": "Unauthorized"}), 401
    records = load_db()
    logs    = load_logs(500)
    today   = datetime.date.today().isoformat()
    granted = [l for l in logs if l.get("status_label") == "GRANTED"]
    denied  = [l for l in logs if l.get("status_label") == "DENIED"]
    today_g = [l for l in granted if l.get("timestamp","").startswith(today)]
    return jsonify({
        "total_students"    : len([r for r in records if r.get("role") == "student"]),
        "total_staff"       : len([r for r in records if r.get("role") == "staff"]),
        "total_admins_db"   : len([r for r in records if r.get("role") == "admin"]),
        "total_scans"       : len(logs),
        "total_granted"     : len(granted),
        "total_denied"      : len(denied),
        "today_granted"     : len(today_g),
        "admins_registered" : len(load_admin_db()),
    })

@app.route("/admin/logs")
def admin_logs():
    if not is_admin_session(request):
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify(load_logs(200))

@app.route("/admin/delete/<roll_id>", methods=["DELETE"])
def admin_delete(roll_id):
    if not is_admin_session(request):
        return jsonify({"error": "Unauthorized"}), 401
    records = load_db()
    records = [r for r in records if r["id"] != roll_id]
    save_db(records)
    return jsonify({"ok": True})

@app.route("/admin/settings", methods=["GET"])
def get_settings_route():
    if not is_admin_session(request):
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify({
        "tolerance"    : TOLERANCE,
        "camera_index" : CAMERA_INDEX,
        "scan_interval": SCAN_INTERVAL
    })

@app.route("/admin/settings", methods=["POST"])
def save_settings_route():
    if not is_admin_session(request):
        return jsonify({"error": "Unauthorized"}), 401
    global TOLERANCE, CAMERA_INDEX, SCAN_INTERVAL
    data = request.json
    if "tolerance"     in data: TOLERANCE     = float(data["tolerance"])
    if "camera_index"  in data: CAMERA_INDEX  = int(data["camera_index"])
    if "scan_interval" in data: SCAN_INTERVAL = float(data["scan_interval"])
    save_settings_file()
    return jsonify({"ok": True})

# ── Start ─────────────────────────────────────────────────────────────────────
@app.route("/qr/<roll_id>")
def get_qr(roll_id):
    """Return QR image for a student."""
    path = os.path.join(QR_DIR, f"{roll_id}.png")
    if os.path.exists(path):
        return send_file(path, mimetype="image/png")
    return jsonify({"error": "QR not found"}), 404


@app.route("/qr_upload", methods=["POST"])
def qr_upload():
    """
    Receive image (base64) from frontend, decode QR using OpenCV.
    Returns: { ok, qr_data } or { ok: False, error }
    """
    data    = request.json
    img_b64 = data.get("image", "")

    if not img_b64:
        return jsonify({"ok": False, "error": "Image nahi mili."})

    try:
        img_data = base64.b64decode(img_b64.split(",")[-1])
        nparr    = np.frombuffer(img_data, np.uint8)
        frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        if frame is None:
            return jsonify({"ok": False, "error": "Image decode nahi hui."})

        # Try OpenCV QR detection
        data_qr, _, _ = QR_DETECTOR.detectAndDecode(frame)

        if data_qr and data_qr.startswith("FACEGATE|"):
            return jsonify({"ok": True, "qr_data": data_qr})

        # Try grayscale + enhanced contrast
        gray     = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        enhanced = cv2.equalizeHist(gray)
        frame_e  = cv2.cvtColor(enhanced, cv2.COLOR_GRAY2BGR)
        data_qr2, _, _ = QR_DETECTOR.detectAndDecode(frame_e)

        if data_qr2 and data_qr2.startswith("FACEGATE|"):
            return jsonify({"ok": True, "qr_data": data_qr2})

        return jsonify({"ok": False, "error": "QR code image mein nahi mila. Seedha aur clear photo upload karo."})

    except Exception as e:
        return jsonify({"ok": False, "error": f"Error: {str(e)}"})

@app.route("/qr_verify", methods=["POST"])
def qr_verify():
    """
    QR code scan se entry — backup for face recognition.
    Frontend sends: { "qr_data": "FACEGATE|CS001|Rahul|student|CSE" }
    """
    data    = request.json
    qr_data = data.get("qr_data", "").strip()

    if not qr_data.startswith("FACEGATE|"):
        return jsonify({"ok": False, "error": "Invalid QR code."})

    parts = qr_data.split("|")
    if len(parts) < 5:
        return jsonify({"ok": False, "error": "QR data incomplete."})

    _, roll_id, name, role, dept = parts[0], parts[1], parts[2], parts[3], parts[4]

    # Verify in database
    records = load_db()
    person  = next((r for r in records if r["id"] == roll_id), None)

    if not person:
        return jsonify({"ok": False, "error": "QR valid but person not in database."})

    # Mark attendance
    newly   = mark_attendance(roll_id, person["name"], person["role"], person.get("dept","N/A"))
    att_pct = get_attendance_percent(roll_id)

    result = {
        "ok"         : True,
        "status"     : "granted",
        "name"       : person["name"],
        "id"         : roll_id,
        "role"       : person["role"],
        "dept"       : person.get("dept","N/A"),
        "att_pct"    : att_pct,
        "newly_marked": newly,
        "img"        : _img_b64(person.get("img_path","")),
        "entry_type" : "QR",
    }

    append_log({
        "timestamp"   : datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "name"        : person["name"],
        "id"          : roll_id,
        "role"        : person["role"],
        "status_label": "GRANTED",
        "confidence"  : "QR",
        "entry_type"  : "QR"
    })

    return jsonify(result)

@app.route("/attendance/monthly")
def monthly_report():
    """Monthly attendance Excel — separate sheet per month."""
    if not EXCEL_OK:
        return jsonify({"error": "openpyxl not installed"}), 500

    year_str = request.args.get("year",  str(datetime.date.today().year))
    month_str= request.args.get("month", str(datetime.date.today().month))
    year, month = int(year_str), int(month_str)

    import calendar
    month_name  = calendar.month_name[month]
    days_in_month = calendar.monthrange(year, month)[1]
    all_days    = [datetime.date(year, month, d).isoformat() for d in range(1, days_in_month+1)]

    att     = load_attendance()
    records = load_db()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    roles = ["student", "staff", "admin"]
    for role in roles:
        role_records = [r for r in records if r.get("role","").lower() == role]
        if not role_records:
            continue

        ws = wb.create_sheet(title=role.capitalize())

        hdr_fill  = PatternFill("solid", fgColor="2D0A4E")
        hdr_font  = Font(bold=True, color="E9D5FF", size=9)
        grn_fill  = PatternFill("solid", fgColor="064E3B")
        red_fill  = PatternFill("solid", fgColor="450A0A")
        alt_fill  = PatternFill("solid", fgColor="160F26")
        ctr       = Alignment(horizontal="center")

        # Title
        total_cols = 4 + days_in_month + 2
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(1,1).value     = f"FaceGate — {role.upper()} Monthly Report  |  {month_name} {year}"
        ws.cell(1,1).font      = Font(bold=True, color="B94FFF", size=13)
        ws.cell(1,1).alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 26

        # Fixed headers
        fixed_hdrs = ["#", "Name", "ID", "Department"]
        for ci, h in enumerate(fixed_hdrs, 1):
            c = ws.cell(2, ci, h)
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = ctr
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 13
        ws.column_dimensions["D"].width = 16

        # Day headers (1–31)
        for di, day in enumerate(all_days, 5):
            d_obj = datetime.date.fromisoformat(day)
            label = f"{d_obj.day}\n{d_obj.strftime('%a')[:2]}"
            c = ws.cell(2, di, f"{d_obj.day} {d_obj.strftime('%a')[:2]}")
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = ctr
            ws.column_dimensions[chr(64+di) if di <= 26 else "A" + chr(64+di-26)].width = 5

        # Summary headers
        sum_col = 5 + days_in_month
        ws.cell(2, sum_col,   "Present").fill = hdr_fill; ws.cell(2, sum_col).font   = hdr_font; ws.cell(2, sum_col).alignment = ctr
        ws.cell(2, sum_col+1, "Att %").fill   = hdr_fill; ws.cell(2, sum_col+1).font = hdr_font; ws.cell(2, sum_col+1).alignment = ctr

        for ri, r in enumerate(role_records, 1):
            sid  = r["id"]
            recs = att.get(sid, {}).get("records", {})
            row  = ri + 2

            ws.cell(row, 1, ri).alignment = ctr
            ws.cell(row, 2, r["name"])
            ws.cell(row, 3, sid).alignment = ctr
            ws.cell(row, 4, r.get("dept","N/A")).alignment = ctr
            if ri % 2 == 0:
                for ci in range(1, 5):
                    ws.cell(row, ci).fill = alt_fill

            present_count = 0
            for di, day in enumerate(all_days, 5):
                c = ws.cell(row, di)
                if day in recs:
                    c.value = "✓"
                    c.fill  = grn_fill
                    c.font  = Font(color="34D399", bold=True)
                    present_count += 1
                else:
                    c.value = "✗"
                    c.fill  = red_fill
                    c.font  = Font(color="F43F5E")
                c.alignment = ctr

            pct = round(present_count / days_in_month * 100, 1)
            pct_font_color = "34D399" if pct >= 75 else "F43F5E"
            ws.cell(row, sum_col,   present_count).alignment = ctr
            ws.cell(row, sum_col,   present_count).font = Font(color="C084FC", bold=True)
            ws.cell(row, sum_col+1, f"{pct}%").alignment = ctr
            ws.cell(row, sum_col+1).font = Font(color=pct_font_color, bold=True)

    fname = f"Monthly_Report_{month_name}_{year}.xlsx"
    path  = os.path.join(BASE_DIR, "data", fname)
    wb.save(path)
    return send_file(path, as_attachment=True, download_name=fname)


if __name__ == "__main__":
    print("\n  ╔══════════════════════════════════════════════╗")
    print("  ║   FaceGate — Updated Server Starting...     ║")
    print("  ║   Open:  http://localhost:5000               ║")
    print("  ║   Admin: http://localhost:5000/admin         ║")
    print("  ╚══════════════════════════════════════════════╝\n")
    if not admin_registered():
        print("  ⚠️  Koi admin register nahi hai!")
        print("  ➜  /admin page par jaake pehle admin register karein.\n")
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)